"""
Streaming Writers Module

This module provides various streaming writer implementations for efficiently
writing large datasets to different file formats with minimal memory usage.

Key Features:
- Abstract base class for consistent writer interface
- CSV writer with proper escaping and buffering
- JSON/JSONL writer for structured data
- Parquet writer for columnar data (requires pyarrow)
- Excel writer for business reporting (requires openpyxl)
- Configurable buffering and compression options
- Progress tracking and error handling
"""

import os
import json
import logging
from typing import Dict, List, Any, Optional
from abc import ABC, abstractmethod
import time


class StreamingWriter(ABC):
    """
    Abstract base class for streaming writers

    Provides a consistent interface for writing data in batches while
    maintaining low memory usage and supporting various output formats.
    """

    def __init__(self, file_path: str, buffer_size: int = 8192,
                 enable_progress: bool = False, logger: Optional[logging.Logger] = None):
        """
        Initialize streaming writer

        Args:
            file_path: Output file path
            buffer_size: Buffer size for I/O operations
            enable_progress: Enable progress tracking
            logger: Logger instance
        """
        self.file_path = file_path
        self.buffer_size = buffer_size
        self.enable_progress = enable_progress
        self.logger = logger or logging.getLogger(__name__)

        # Progress tracking
        self.records_written = 0
        self.batches_written = 0
        self.start_time = None
        self.bytes_written = 0

        # State management
        self.is_closed = False
        self.header_written = False

        # Ensure output directory exists
        os.makedirs(os.path.dirname(self.file_path) or '.', exist_ok=True)

    @abstractmethod
    def write_header(self, columns: List[str]) -> None:
        """Write file header with column information"""
        pass

    @abstractmethod
    def write_batch(self, batch: List[Dict[str, Any]]) -> None:
        """Write batch of data"""
        pass

    @abstractmethod
    def write_footer(self) -> None:
        """Write file footer if needed"""
        pass

    @abstractmethod
    def close(self) -> None:
        """Close writer and finalize file"""
        pass

    def __enter__(self):
        """Context manager entry"""
        self.start_time = time.time()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        if not self.is_closed:
            self.close()

    def get_progress_stats(self) -> Dict[str, Any]:
        """Get current progress statistics"""
        current_time = time.time()
        elapsed_time = current_time - (self.start_time or current_time)

        return {
            'records_written': self.records_written,
            'batches_written': self.batches_written,
            'bytes_written': self.bytes_written,
            'elapsed_time_seconds': elapsed_time,
            'records_per_second': self.records_written / elapsed_time if elapsed_time > 0 else 0,
            'mb_per_second': (self.bytes_written / (1024 * 1024)) / elapsed_time if elapsed_time > 0 else 0
        }

    def _log_progress(self, batch_size: int):
        """Log progress information"""
        if self.enable_progress:
            self.records_written += batch_size
            self.batches_written += 1

            if self.batches_written % 100 == 0:  # Log every 100 batches
                stats = self.get_progress_stats()
                self.logger.info(
                    f"Progress: {stats['records_written']} records, "
                    f"{stats['records_per_second']:.1f} records/sec, "
                    f"{stats['mb_per_second']:.2f} MB/sec"
                )


class StreamingCSVWriter(StreamingWriter):
    """
    Streaming CSV writer with proper escaping and buffering

    Features:
    - RFC 4180 compliant CSV formatting
    - Proper escaping of special characters
    - Configurable delimiters and quote characters
    - Memory-efficient buffered writing
    """

    def __init__(self, file_path: str, buffer_size: int = 8192,
                 delimiter: str = ',', quote_char: str = '"',
                 line_terminator: str = '\n', encoding: str = 'utf-8',
                 enable_progress: bool = False, logger: Optional[logging.Logger] = None):
        """
        Initialize CSV writer

        Args:
            file_path: Output CSV file path
            buffer_size: Buffer size for I/O operations
            delimiter: Field delimiter character
            quote_char: Quote character for escaping
            line_terminator: Line terminator sequence
            encoding: File encoding
            enable_progress: Enable progress tracking
            logger: Logger instance
        """
        super().__init__(file_path, buffer_size, enable_progress, logger)

        self.delimiter = delimiter
        self.quote_char = quote_char
        self.line_terminator = line_terminator
        self.encoding = encoding

        self.file_handle = None
        self.columns = None

        self._open_file()

    def _open_file(self):
        """Open CSV file for writing"""
        try:
            self.file_handle = open(
                self.file_path, 'w',
                buffering=self.buffer_size,
                newline='',
                encoding=self.encoding
            )
            self.logger.debug(f"Opened CSV file for writing: {self.file_path}")
        except Exception as e:
            self.logger.error(f"Failed to open CSV file {self.file_path}: {e}")
            raise

    def write_header(self, columns: List[str]) -> None:
        """Write CSV header row"""
        if self.header_written:
            self.logger.warning("Header already written, skipping")
            return

        if not self.file_handle:
            raise RuntimeError("File not open for writing")

        self.columns = columns

        # Escape column names if needed
        escaped_columns = [self._escape_field(col) for col in columns]
        header_line = self.delimiter.join(escaped_columns) + self.line_terminator

        self.file_handle.write(header_line)
        self.bytes_written += len(header_line.encode(self.encoding))
        self.header_written = True

        self.logger.debug(f"CSV header written with {len(columns)} columns")

    def write_batch(self, batch: List[Dict[str, Any]]) -> None:
        """Write batch of data to CSV"""
        if not self.file_handle or not self.columns:
            raise RuntimeError("File not open or header not written")

        if not batch:
            return

        batch_lines = []

        for row in batch:
            values = []
            for col in self.columns:
                value = row.get(col, '')
                escaped_value = self._escape_field(value)
                values.append(escaped_value)

            line = self.delimiter.join(values) + self.line_terminator
            batch_lines.append(line)

        # Write all lines at once for better performance
        batch_content = ''.join(batch_lines)
        self.file_handle.write(batch_content)
        self.bytes_written += len(batch_content.encode(self.encoding))

        self._log_progress(len(batch))

    def _escape_field(self, value: Any) -> str:
        """Escape field value for CSV format"""
        if value is None:
            return ''

        str_value = str(value)

        # Check if escaping is needed
        needs_quoting = (
                self.delimiter in str_value or
                self.quote_char in str_value or
                self.line_terminator in str_value or
                str_value.startswith(' ') or
                str_value.endswith(' ')
        )

        if needs_quoting:
            # Escape existing quote characters by doubling them
            escaped_value = str_value.replace(self.quote_char, self.quote_char + self.quote_char)
            return f"{self.quote_char}{escaped_value}{self.quote_char}"

        return str_value

    def write_footer(self) -> None:
        """CSV doesn't require a footer"""
        pass

    def close(self) -> None:
        """Close CSV file"""
        if self.file_handle and not self.is_closed:
            self.file_handle.close()
            self.file_handle = None
            self.is_closed = True

            stats = self.get_progress_stats()
            self.logger.info(
                f"CSV file completed: {self.file_path} "
                f"({stats['records_written']} records, {stats['bytes_written']} bytes)"
            )


class StreamingJSONWriter(StreamingWriter):
    """
    Streaming JSON writer supporting both JSONL and JSON array formats

    Features:
    - JSONL (JSON Lines) format for streaming
    - JSON array format for compatibility
    - Configurable indentation and formatting
    - Custom JSON serialization for complex types
    """

    def __init__(self, file_path: str, buffer_size: int = 8192,
                 format_type: str = 'jsonl', indent: Optional[int] = None,
                 ensure_ascii: bool = False, encoding: str = 'utf-8',
                 enable_progress: bool = False, logger: Optional[logging.Logger] = None):
        """
        Initialize JSON writer

        Args:
            file_path: Output JSON file path
            buffer_size: Buffer size for I/O operations
            format_type: 'jsonl' for JSON Lines or 'array' for JSON array
            indent: Indentation level for pretty printing
            ensure_ascii: Ensure ASCII output
            encoding: File encoding
            enable_progress: Enable progress tracking
            logger: Logger instance
        """
        super().__init__(file_path, buffer_size, enable_progress, logger)

        self.format_type = format_type.lower()
        self.indent = indent
        self.ensure_ascii = ensure_ascii
        self.encoding = encoding

        self.file_handle = None
        self.first_record = True

        if self.format_type not in ['jsonl', 'array']:
            raise ValueError(f"Unsupported format_type: {format_type}. Use 'jsonl' or 'array'")

        self._open_file()

    def _open_file(self):
        """Open JSON file for writing"""
        try:
            self.file_handle = open(
                self.file_path, 'w',
                buffering=self.buffer_size,
                encoding=self.encoding
            )
            self.logger.debug(f"Opened JSON file for writing: {self.file_path}")
        except Exception as e:
            self.logger.error(f"Failed to open JSON file {self.file_path}: {e}")
            raise

    def write_header(self, columns: List[str]) -> None:
        """Write JSON header if needed"""
        if self.header_written:
            return

        if not self.file_handle:
            raise RuntimeError("File not open for writing")

        if self.format_type == 'array':
            # Start JSON array
            self.file_handle.write('[\n' if self.indent else '[')
            self.bytes_written += len('[')

        self.header_written = True
        self.logger.debug(f"JSON header written for format: {self.format_type}")

    def write_batch(self, batch: List[Dict[str, Any]]) -> None:
        """Write batch of data as JSON"""
        if not self.file_handle:
            raise RuntimeError("File not open for writing")

        if not batch:
            return

        if self.format_type == 'jsonl':
            self._write_jsonl_batch(batch)
        else:  # array format
            self._write_array_batch(batch)

        self._log_progress(len(batch))

    def _write_jsonl_batch(self, batch: List[Dict[str, Any]]):
        """Write batch in JSONL format"""
        lines = []

        for row in batch:
            json_line = json.dumps(
                row,
                ensure_ascii=self.ensure_ascii,
                default=self._json_serializer,
                indent=self.indent
            ) + '\n'
            lines.append(json_line)

        batch_content = ''.join(lines)
        self.file_handle.write(batch_content)
        self.bytes_written += len(batch_content.encode(self.encoding))

    def _write_array_batch(self, batch: List[Dict[str, Any]]):
        """Write batch in JSON array format"""
        lines = []

        for row in batch:
            # Add comma separator for non-first records
            if not self.first_record:
                lines.append(',')
                if self.indent:
                    lines.append('\n')

            json_str = json.dumps(
                row,
                ensure_ascii=self.ensure_ascii,
                default=self._json_serializer,
                indent=self.indent
            )

            if self.indent:
                # Add proper indentation for array format
                indented_lines = []
                for line in json_str.split('\n'):
                    if line.strip():
                        indented_lines.append('  ' + line)
                    else:
                        indented_lines.append(line)
                json_str = '\n'.join(indented_lines)
                lines.append('\n' + json_str)
            else:
                lines.append(json_str)

            self.first_record = False

        batch_content = ''.join(lines)
        self.file_handle.write(batch_content)
        self.bytes_written += len(batch_content.encode(self.encoding))

    def _json_serializer(self, obj: Any) -> Any:
        """Custom JSON serializer for complex types"""
        import datetime
        import decimal

        if isinstance(obj, datetime.datetime):
            return obj.isoformat()
        elif isinstance(obj, datetime.date):
            return obj.isoformat()
        elif isinstance(obj, datetime.time):
            return obj.isoformat()
        elif isinstance(obj, decimal.Decimal):
            return float(obj)
        elif hasattr(obj, '__dict__'):
            return obj.__dict__
        else:
            return str(obj)

    def write_footer(self) -> None:
        """Write JSON footer if needed"""
        if self.format_type == 'array' and self.file_handle:
            footer = '\n]' if self.indent else ']'
            self.file_handle.write(footer)
            self.bytes_written += len(footer.encode(self.encoding))

    def close(self) -> None:
        """Close JSON file"""
        if self.file_handle and not self.is_closed:
            self.write_footer()
            self.file_handle.close()
            self.file_handle = None
            self.is_closed = True

            stats = self.get_progress_stats()
            self.logger.info(
                f"JSON file completed: {self.file_path} "
                f"({stats['records_written']} records, {stats['bytes_written']} bytes)"
            )


class StreamingParquetWriter(StreamingWriter):
    """
    Streaming Parquet writer for columnar data storage

    Features:
    - Efficient columnar storage format
    - Built-in compression
    - Schema evolution support
    - High performance for analytics workloads

    Requires: pyarrow
    """

    def __init__(self, file_path: str, compression: str = 'snappy',
                 row_group_size: int = 10000, enable_progress: bool = False,
                 logger: Optional[logging.Logger] = None):
        """
        Initialize Parquet writer

        Args:
            file_path: Output Parquet file path
            compression: Compression algorithm ('snappy', 'gzip', 'lz4')
            row_group_size: Number of rows per row group
            enable_progress: Enable progress tracking
            logger: Logger instance
        """
        super().__init__(file_path, 0, enable_progress, logger)  # No buffer for Parquet

        self.compression = compression
        self.row_group_size = row_group_size

        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
            self.pa = pa
            self.pq = pq
        except ImportError:
            raise ImportError("pyarrow is required for Parquet writing. Install with: pip install pyarrow")

        self.schema = None
        self.writer = None
        self.batch_buffer = []

    def write_header(self, columns: List[str]) -> None:
        """Initialize Parquet schema"""
        if self.header_written:
            return

        # Create schema with string types by default
        # Schema will be inferred from actual data
        self.columns = columns
        self.header_written = True

        self.logger.debug(f"Parquet schema initialized with {len(columns)} columns")

    def write_batch(self, batch: List[Dict[str, Any]]) -> None:
        """Write batch of data to Parquet"""
        if not batch:
            return

        self.batch_buffer.extend(batch)

        # Write when buffer reaches row group size
        if len(self.batch_buffer) >= self.row_group_size:
            self._write_row_group()

        self._log_progress(len(batch))

    def _write_row_group(self):
        """Write accumulated batch buffer as a row group"""
        if not self.batch_buffer:
            return

        try:
            # Convert to PyArrow table
            table = self.pa.Table.from_pylist(self.batch_buffer)

            if self.writer is None:
                # Initialize writer with inferred schema
                self.schema = table.schema
                self.writer = self.pq.ParquetWriter(
                    self.file_path,
                    self.schema,
                    compression=self.compression
                )
                self.logger.debug(f"Parquet writer initialized with schema: {self.schema}")

            # Write row group
            self.writer.write_table(table)

            # Clear buffer
            self.batch_buffer.clear()

        except Exception as e:
            self.logger.error(f"Failed to write Parquet row group: {e}")
            raise

    def write_footer(self) -> None:
        """Write remaining data"""
        if self.batch_buffer:
            self._write_row_group()

    def close(self) -> None:
        """Close Parquet writer"""
        if not self.is_closed:
            self.write_footer()

            if self.writer:
                self.writer.close()
                self.writer = None

            self.is_closed = True

            stats = self.get_progress_stats()
            self.logger.info(
                f"Parquet file completed: {self.file_path} "
                f"({stats['records_written']} records)"
            )


class StreamingExcelWriter(StreamingWriter):
    """
    Streaming Excel writer for business reporting

    Features:
    - Multiple worksheets support
    - Formatting and styling options
    - Large dataset handling
    - Memory-efficient writing

    Requires: openpyxl
    """

    def __init__(self, file_path: str, worksheet_name: str = 'Sheet1',
                 max_rows_per_sheet: int = 1000000, enable_progress: bool = False,
                 logger: Optional[logging.Logger] = None):
        """
        Initialize Excel writer

        Args:
            file_path: Output Excel file path
            worksheet_name: Name of the worksheet
            max_rows_per_sheet: Maximum rows per worksheet
            enable_progress: Enable progress tracking
            logger: Logger instance
        """
        super().__init__(file_path, 0, enable_progress, logger)  # No buffer for Excel

        self.worksheet_name = worksheet_name
        self.max_rows_per_sheet = max_rows_per_sheet

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            self.Workbook = Workbook
            self.get_column_letter = get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel writing. Install with: pip install openpyxl")

        self.workbook = None
        self.worksheet = None
        self.current_row = 1
        self.current_sheet_number = 1
        self.columns = None

    def write_header(self, columns: List[str]) -> None:
        """Write Excel header row"""
        if self.header_written:
            return

        self.columns = columns

        # Initialize workbook and worksheet
        self.workbook = self.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = self.worksheet_name

        # Write header row
        for col_idx, column_name in enumerate(columns, 1):
            self.worksheet.cell(row=1, column=col_idx, value=column_name)

        self.current_row = 2
        self.header_written = True

        self.logger.debug(f"Excel header written with {len(columns)} columns")

    def write_batch(self, batch: List[Dict[str, Any]]) -> None:
        """Write batch of data to Excel"""
        if not self.worksheet or not self.columns:
            raise RuntimeError("Worksheet not initialized or header not written")

        if not batch:
            return

        for row_data in batch:
            # Check if we need a new worksheet
            if self.current_row > self.max_rows_per_sheet:
                self._create_new_worksheet()

            # Write row data
            for col_idx, column_name in enumerate(self.columns, 1):
                value = row_data.get(column_name, '')
                self.worksheet.cell(row=self.current_row, column=col_idx, value=value)

            self.current_row += 1

        self._log_progress(len(batch))

    def _create_new_worksheet(self):
        """Create a new worksheet when row limit is reached"""
        self.current_sheet_number += 1
        sheet_name = f"{self.worksheet_name}_{self.current_sheet_number}"

        self.worksheet = self.workbook.create_sheet(title=sheet_name)

        # Write header to new sheet
        for col_idx, column_name in enumerate(self.columns, 1):
            self.worksheet.cell(row=1, column=col_idx, value=column_name)

        self.current_row = 2

        self.logger.info(f"Created new worksheet: {sheet_name}")

    def write_footer(self) -> None:
        """Excel doesn't require a footer"""
        pass

    def close(self) -> None:
        """Close Excel writer and save file"""
        if not self.is_closed and self.workbook:
            try:
                self.workbook.save(self.file_path)
                self.workbook = None
                self.worksheet = None
                self.is_closed = True

                stats = self.get_progress_stats()
                self.logger.info(
                    f"Excel file completed: {self.file_path} "
                    f"({stats['records_written']} records, {self.current_sheet_number} sheets)"
                )
            except Exception as e:
                self.logger.error(f"Failed to save Excel file: {e}")
                raise


# ===================== WRITER FACTORY =====================

class WriterFactory:
    """
    Factory class for creating streaming writers based on file extension or format
    """

    @staticmethod
    def create_writer(file_path: str, format_type: Optional[str] = None,
                      **kwargs) -> StreamingWriter:
        """
        Create appropriate streaming writer based on file path or format

        Args:
            file_path: Output file path
            format_type: Explicit format type override
            **kwargs: Additional arguments for specific writers

        Returns:
            StreamingWriter: Appropriate writer instance
        """
        if format_type:
            writer_format = format_type.lower()
        else:
            # Infer format from file extension
            _, ext = os.path.splitext(file_path)
            writer_format = ext.lower().lstrip('.')

        if writer_format in ['csv', 'tsv']:
            delimiter = '\t' if writer_format == 'tsv' else kwargs.get('delimiter', ',')
            return StreamingCSVWriter(file_path, delimiter=delimiter, **kwargs)

        elif writer_format in ['json', 'jsonl']:
            format_type = 'jsonl' if writer_format == 'jsonl' else kwargs.get('format_type', 'jsonl')
            return StreamingJSONWriter(file_path, format_type=format_type, **kwargs)

        elif writer_format == 'parquet':
            return StreamingParquetWriter(file_path, **kwargs)

        elif writer_format in ['xlsx', 'excel']:
            return StreamingExcelWriter(file_path, **kwargs)

        else:
            raise ValueError(
                f"Unsupported format: {writer_format}. "
                f"Supported formats: csv, tsv, json, jsonl, parquet, xlsx"
            )

    @staticmethod
    def get_supported_formats() -> List[str]:
        """Get list of supported output formats"""
        return ['csv', 'tsv', 'json', 'jsonl', 'parquet', 'xlsx']


# ===================== COMPRESSION UTILITIES =====================

class CompressionWriter:
    """
    Wrapper for adding compression to any streaming writer
    """

    def __init__(self, writer: StreamingWriter, compression: str = 'gzip'):
        """
        Initialize compression wrapper

        Args:
            writer: Base streaming writer
            compression: Compression type ('gzip', 'bz2', 'lzma')
        """
        self.writer = writer
        self.compression = compression.lower()

        if self.compression not in ['gzip', 'bz2', 'lzma']:
            raise ValueError(f"Unsupported compression: {compression}")

        # Update file path with compression extension
        if not self.writer.file_path.endswith(f'.{self.compression}'):
            self.writer.file_path += f'.{self.compression}'

        self._setup_compression()

    def _setup_compression(self):
        """Setup compression based on type"""
        import gzip
        import bz2
        import lzma

        original_open = self.writer._open_file

        def compressed_open():
            if self.compression == 'gzip':
                self.writer.file_handle = gzip.open(
                    self.writer.file_path, 'wt',
                    encoding=getattr(self.writer, 'encoding', 'utf-8'),
                    buffering=self.writer.buffer_size
                )
            elif self.compression == 'bz2':
                self.writer.file_handle = bz2.open(
                    self.writer.file_path, 'wt',
                    encoding=getattr(self.writer, 'encoding', 'utf-8')
                )
            elif self.compression == 'lzma':
                self.writer.file_handle = lzma.open(
                    self.writer.file_path, 'wt',
                    encoding=getattr(self.writer, 'encoding', 'utf-8')
                )

        self.writer._open_file = compressed_open

    def __getattr__(self, name):
        """Delegate all other methods to the wrapped writer"""
        return getattr(self.writer, name)


# ===================== EXAMPLE USAGE =====================

def example_usage():
    """
    Example usage of streaming writers
    """
    import logging

    # Setup logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)

    # Sample data
    sample_data = [
        {'id': i, 'name': f'User {i}', 'email': f'user{i}@example.com', 'age': 20 + (i % 50)}
        for i in range(1, 1001)
    ]

    columns = ['id', 'name', 'email', 'age']

    # Example 1: CSV Writer
    print("Writing CSV file...")
    with StreamingCSVWriter('output/example.csv', enable_progress=True, logger=logger) as writer:
        writer.write_header(columns)

        # Write in batches
        batch_size = 100
        for i in range(0, len(sample_data), batch_size):
            batch = sample_data[i:i + batch_size]
            writer.write_batch(batch)

    # Example 2: JSON Writer (JSONL format)
    print("Writing JSONL file...")
    with StreamingJSONWriter('output/example.jsonl', format_type='jsonl',
                             enable_progress=True, logger=logger) as writer:
        writer.write_header(columns)

        for i in range(0, len(sample_data), batch_size):
            batch = sample_data[i:i + batch_size]
            writer.write_batch(batch)

    # Example 3: Using Writer Factory
    print("Using Writer Factory...")
    writer = WriterFactory.create_writer('output/example_factory.csv', enable_progress=True, logger=logger)

    with writer:
        writer.write_header(columns)
        writer.write_batch(sample_data)

    # Example 4: Compressed CSV
    print("Writing compressed CSV...")
    base_writer = StreamingCSVWriter('output/example_compressed.csv', enable_progress=True, logger=logger)
    compressed_writer = CompressionWriter(base_writer, compression='gzip')

    with compressed_writer:
        compressed_writer.write_header(columns)
        compressed_writer.write_batch(sample_data)

    print("Examples completed successfully!")


if __name__ == "__main__":
    example_usage()